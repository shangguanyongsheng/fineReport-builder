"""Excel 文件解析器 - 辅助生成 .cpt 报表

将 Excel 模板转换为帆软报表结构：
- 单元格布局 → CellElementList
- 样式 → StyleList
- 筛选区域 → ReportParameterAttr
- 数据透视表 → TableDataMap + SQL
"""
import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import json


@dataclass
class ExcelCell:
    """Excel 单元格"""
    column: int  # 0-indexed
    row: int     # 0-indexed
    value: Any
    formula: str = ""
    merge_start: Tuple[int, int] = None  # 合并单元格起始 (col, row)
    merge_end: Tuple[int, int] = None    # 合并单元格结束
    style_index: int = 0
    data_type: str = "text"  # text, number, date, formula


@dataclass  
class ExcelStyle:
    """Excel 样式"""
    font_name: str = "SimSun"
    font_size: int = 11
    bold: bool = False
    italic: bool = False
    font_color: str = "#000000"
    bg_color: str = "#FFFFFF"
    horizontal: str = "left"  # left, center, right
    vertical: str = "center"  # top, center, bottom
    border_top: str = ""
    border_bottom: str = ""
    border_left: str = ""
    border_right: str = ""
    number_format: str = ""


@dataclass
class ExcelSheet:
    """Excel 工作表"""
    name: str
    cells: List[ExcelCell] = field(default_factory=list)
    styles: List[ExcelStyle] = field(default_factory=list)
    merged_ranges: List[Tuple[int, int, int, int]] = field(default_factory=list)
    column_widths: Dict[int, int] = field(default_factory=dict)
    row_heights: Dict[int, int] = field(default_factory=dict)
    filter_range: Optional[str] = None  # 筛选区域
    max_column: int = 0
    max_row: int = 0


@dataclass
class ExcelStructure:
    """Excel 文件结构"""
    sheets: List[ExcelSheet] = field(default_factory=list)
    defined_names: Dict[str, str] = field(default_factory=dict)  # 名称管理器


class ExcelParser:
    """Excel 文件解析器"""
    
    def parse(self, file_path: str) -> ExcelStructure:
        """解析 Excel 文件"""
        wb = openpyxl.load_workbook(file_path, data_only=False)
        structure = ExcelStructure()
        
        # 解析每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = ExcelSheet(name=sheet_name)
            
            # 解析合并单元格
            for merged_range in ws.merged_cells.ranges:
                sheet.merged_ranges.append((
                    merged_range.min_col - 1,  # 转为 0-indexed
                    merged_range.min_row - 1,
                    merged_range.max_col - 1,
                    merged_range.max_row - 1
                ))
            
            # 解析单元格
            style_map = {}  # 样式缓存
            for row_idx, row in enumerate(ws.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    excel_cell = self._parse_cell(cell, col_idx - 1, row_idx - 1, style_map)
                    sheet.cells.append(excel_cell)
                    
                    sheet.max_column = max(sheet.max_column, col_idx - 1)
                    sheet.max_row = max(sheet.max_row, row_idx - 1)
            
            # 转换样式映射为列表
            sheet.styles = list(style_map.values())
            
            # 为单元格设置样式索引
            style_list = [(s.font_name, s.font_size, s.bold, s.bg_color, s.horizontal) 
                         for s in sheet.styles]
            for cell in sheet.cells:
                cell_style = self._get_cell_style(ws, cell.column + 1, cell.row + 1)
                style_key = (
                    cell_style.font_name,
                    cell_style.font_size, 
                    cell_style.bold,
                    cell_style.bg_color,
                    cell_style.horizontal
                )
                if style_key in style_list:
                    cell.style_index = style_list.index(style_key)
            
            # 解析列宽
            for col_idx, col_dim in ws.column_dimensions.items():
                col_num = openpyxl.utils.column_index_from_string(col_idx) - 1
                sheet.column_widths[col_num] = int(col_dim.width or 10)
            
            # 解析行高
            for row_idx, row_dim in ws.row_dimensions.items():
                sheet.row_heights[row_idx - 1] = int(row_dim.height or 20)
            
            # 解析筛选区域
            if ws.auto_filter.ref:
                sheet.filter_range = ws.auto_filter.ref
            
            structure.sheets.append(sheet)
        
        # 解析名称管理器
        for name in wb.defined_names:
            defined_name = wb.defined_names[name]
            structure.defined_names[name] = defined_name.attr_text if hasattr(defined_name, 'attr_text') else str(defined_name)
        
        wb.close()
        return structure
    
    def _parse_cell(self, cell, col: int, row: int, style_map: Dict) -> ExcelCell:
        """解析单个单元格"""
        excel_cell = ExcelCell(column=col, row=row, value=cell.value)
        
        # 判断数据类型
        if cell.data_type == 'f':
            excel_cell.data_type = "formula"
            excel_cell.formula = cell.value
        elif cell.data_type == 'n':
            excel_cell.data_type = "number"
        elif cell.data_type == 'd':
            excel_cell.data_type = "date"
        else:
            excel_cell.data_type = "text"
        
        # 检查是否是合并单元格的一部分
        if cell.coordinate in cell.parent.merged_cells:
            for merged_range in cell.parent.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    excel_cell.merge_start = (merged_range.min_col - 1, merged_range.min_row - 1)
                    excel_cell.merge_end = (merged_range.max_col - 1, merged_range.max_row - 1)
                    break
        
        return excel_cell
    
    def _get_cell_style(self, ws, col: int, row: int) -> ExcelStyle:
        """获取单元格样式"""
        cell = ws.cell(row=row, column=col)
        style = ExcelStyle()
        
        # 字体
        if cell.font:
            style.font_name = cell.font.name or "SimSun"
            style.font_size = cell.font.size or 11
            style.bold = cell.font.bold or False
            style.italic = cell.font.italic or False
            if cell.font.color and cell.font.color.rgb:
                style.font_color = f"#{cell.font.color.rgb}"
        
        # 填充
        if cell.fill and isinstance(cell.fill, PatternFill):
            if cell.fill.fgColor and cell.fill.fgColor.rgb:
                style.bg_color = f"#{cell.fill.fgColor.rgb}"
        
        # 对齐
        if cell.alignment:
            style.horizontal = cell.alignment.horizontal or "left"
            style.vertical = cell.alignment.vertical or "center"
        
        # 数字格式
        if cell.number_format:
            style.number_format = cell.number_format
        
        return style
    
    def to_cpt_cells(self, sheet: ExcelSheet) -> List[Dict]:
        """转换为 .cpt 单元格格式"""
        cells = []
        
        for cell in sheet.cells:
            cpt_cell = {
                "column": cell.column,
                "row": cell.row,
                "col_span": 1,
                "row_span": 1,
                "value": str(cell.value) if cell.value is not None else "",
                "value_type": cell.data_type,
                "style_index": cell.style_index
            }
            
            # 处理合并单元格
            if cell.merge_start and cell.merge_start == (cell.column, cell.row):
                cpt_cell["col_span"] = cell.merge_end[0] - cell.merge_start[0] + 1
                cpt_cell["row_span"] = cell.merge_end[1] - cell.merge_start[1] + 1
            
            # 处理公式
            if cell.formula:
                cpt_cell["value"] = f"={cell.formula}"
                cpt_cell["value_type"] = "formula"
            
            cells.append(cpt_cell)
        
        return cells
    
    def to_cpt_styles(self, styles: List[ExcelStyle]) -> List[Dict]:
        """转换为 .cpt 样式格式"""
        cpt_styles = []
        
        for style in styles:
            cpt_style = {
                "font": {
                    "name": style.font_name,
                    "size": style.font_size,
                    "bold": style.bold,
                    "italic": style.italic,
                    "color": style.font_color
                },
                "background": style.bg_color,
                "alignment": {
                    "horizontal": style.horizontal,
                    "vertical": style.vertical
                },
                "border": {}
            }
            
            if style.border_top:
                cpt_style["border"]["top"] = {"style": "thin", "color": "#000000"}
            if style.border_bottom:
                cpt_style["border"]["bottom"] = {"style": "thin", "color": "#000000"}
            if style.border_left:
                cpt_style["border"]["left"] = {"style": "thin", "color": "#000000"}
            if style.border_right:
                cpt_style["border"]["right"] = {"style": "thin", "color": "#000000"}
            
            cpt_styles.append(cpt_style)
        
        return cpt_styles
    
    def generate_preview_html(self, structure: ExcelStructure, sheet_index: int = 0) -> str:
        """生成预览 HTML"""
        if sheet_index >= len(structure.sheets):
            return "<p>No sheets found</p>"
        
        sheet = structure.sheets[sheet_index]
        styles = self.to_cpt_styles(sheet.styles)
        
        html_parts = ['''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel 预览 → .cpt 转换</title>
    <style>
        * { box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 1400px; margin: 0 auto; }
        h1 { color: #333; margin-bottom: 20px; }
        .toolbar { display: flex; gap: 12px; margin-bottom: 16px; padding: 16px; background: white; border-radius: 8px; }
        .btn { padding: 8px 16px; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; }
        .btn-primary { background: #1890ff; color: white; }
        .btn-secondary { background: #f0f0f0; color: #333; }
        .preview-container { background: white; border-radius: 8px; overflow: hidden; }
        .sheet-tabs { display: flex; background: #fafafa; border-bottom: 1px solid #eee; }
        .sheet-tab { padding: 12px 20px; cursor: pointer; border-bottom: 2px solid transparent; }
        .sheet-tab.active { background: white; border-bottom-color: #1890ff; }
        .sheet-tab:hover { background: #e6f7ff; }
        .table-wrapper { overflow: auto; max-height: 70vh; }
        table { border-collapse: collapse; width: 100%; }
        th, td { border: 1px solid #d9d9d9; padding: 8px 12px; min-width: 80px; height: 28px; }
        th { background: #fafafa; font-weight: 500; }
        .cell-formula { color: #722ed1; font-family: monospace; }
        .cell-number { text-align: right; color: #1890ff; }
        .cell-header { background: #e6f7ff; font-weight: bold; }
        .config-panel { position: fixed; right: 20px; top: 80px; width: 320px; background: white; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.15); padding: 16px; }
        .config-panel h3 { margin: 0 0 12px 0; font-size: 14px; color: #666; }
        .config-section { margin-bottom: 16px; }
        .config-section label { display: block; margin-bottom: 4px; font-size: 12px; color: #666; }
        .config-section input, .config-section select { width: 100%; padding: 6px 8px; border: 1px solid #d9d9d9; border-radius: 4px; }
        .field-row { display: flex; align-items: center; gap: 8px; padding: 8px 0; border-bottom: 1px solid #f0f0f0; }
        .field-row input { flex: 1; }
        .field-row select { width: 100px; }
        .status-bar { padding: 8px 16px; background: #fafafa; font-size: 12px; color: #666; }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Excel → .cpt 转换工具</h1>
        
        <div class="toolbar">
            <button class="btn btn-primary" onclick="generateCpt()">生成 .cpt 文件</button>
            <button class="btn btn-secondary" onclick="detectDataRegion()">自动检测数据区域</button>
            <button class="btn btn-secondary" onclick="detectHeaders()">识别表头</button>
            <button class="btn btn-secondary" onclick="exportJson()">导出 JSON</button>
        </div>
''']
        
        # 工作表标签
        html_parts.append('        <div class="preview-container">\n            <div class="sheet-tabs">')
        for i, s in enumerate(structure.sheets):
            active = " active" if i == sheet_index else ""
            html_parts.append(f'                <div class="sheet-tab{active}" onclick="switchSheet({i})">{s.name}</div>')
        html_parts.append('            </div>')
        
        # 表格预览
        html_parts.append('            <div class="table-wrapper">\n                <table>')
        
        # 构建表格
        cells_by_pos = {(c.column, c.row): c for c in sheet.cells}
        processed = set()
        
        for row in range(sheet.max_row + 1):
            html_parts.append('                    <tr>')
            for col in range(sheet.max_column + 1):
                if (col, row) in processed:
                    continue
                
                cell = cells_by_pos.get((col, row))
                if not cell:
                    html_parts.append(f'                        <td></td>')
                    continue
                
                # 检查合并
                rowspan = 1
                colspan = 1
                if cell.merge_start and cell.merge_start == (col, row):
                    colspan = cell.merge_end[0] - cell.merge_start[0] + 1
                    rowspan = cell.merge_end[1] - cell.merge_start[1] + 1
                    # 标记已处理
                    for r in range(row, row + rowspan):
                        for c in range(col, col + colspan):
                            processed.add((c, r))
                elif cell.merge_start:
                    # 被合并的单元格，跳过
                    continue
                
                # 样式类
                style_class = ""
                if cell.data_type == "formula":
                    style_class = "cell-formula"
                elif cell.data_type == "number":
                    style_class = "cell-number"
                
                # 检查是否是表头（第一行或第一列）
                if row == 0 or col == 0:
                    style_class += " cell-header"
                
                attrs = ""
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
                
                value = str(cell.value) if cell.value is not None else ""
                if cell.formula:
                    value = f"={cell.formula}"
                
                tag = "th" if row == 0 else "td"
                html_parts.append(f'                        <{tag}{attrs} class="{style_class}" data-col="{col}" data-row="{row}">{value}</{tag}>')
            
            html_parts.append('                    </tr>')
        
        html_parts.append('                </table>\n            </div>')
        
        # 状态栏
        html_parts.append(f'''
            <div class="status-bar">
                工作表: {sheet.name} | 
                行数: {sheet.max_row + 1} | 
                列数: {sheet.max_column + 1} | 
                样式数: {len(sheet.styles)} |
                合并区域: {len(sheet.merged_ranges)}
            </div>
        </div>
    </div>
''')
        
        # 配置面板
        html_parts.append('''
    <div class="config-panel">
        <h3>⚙️ 报表配置</h3>
        
        <div class="config-section">
            <label>数据源名称</label>
            <input type="text" id="dsName" placeholder="例如: sales_data">
        </div>
        
        <div class="config-section">
            <label>数据库连接</label>
            <select id="dbConnection">
                <option value="">选择数据库</option>
                <option value="cfs-report">cfs-report</option>
                <option value="sales_db">sales_db</option>
            </select>
        </div>
        
        <div class="config-section">
            <label>数据区域（起始行-结束行）</label>
            <div style="display: flex; gap: 8px;">
                <input type="number" id="dataStartRow" placeholder="起始行" value="1">
                <input type="number" id="dataEndRow" placeholder="结束行">
            </div>
        </div>
        
        <div class="config-section">
            <label>字段映射（列 → 数据库字段）</label>
            <div id="fieldMapping">
                <!-- 动态生成 -->
            </div>
        </div>
        
        <button class="btn btn-primary" style="width: 100%; margin-top: 12px;" onclick="applyConfig()">应用配置</button>
    </div>
''')
        
        # JavaScript
        html_parts.append(f'''
    <script>
        const excelData = {json.dumps({
            "sheets": [{"name": s.name, "maxRow": s.max_row, "maxColumn": s.max_column} for s in structure.sheets],
            "currentSheet": sheet_index
        })};
        
        function switchSheet(index) {{
            location.href = '?sheet=' + index;
        }}
        
        function detectDataRegion() {{
            alert('检测数据区域功能 - 需要后端支持');
        }}
        
        function detectHeaders() {{
            // 高亮第一行
            const headers = document.querySelectorAll('tr:first-child th, tr:first-child td');
            headers.forEach(th => th.style.background = '#bae7ff');
            alert('已识别第一行为表头');
        }}
        
        function generateCpt() {{
            const config = {{
                dataSource: document.getElementById('dsName').value,
                database: document.getElementById('dbConnection').value,
                dataStartRow: document.getElementById('dataStartRow').value,
                dataEndRow: document.getElementById('dataEndRow').value
            }};
            
            if (!config.dataSource) {{
                alert('请填写数据源名称');
                return;
            }}
            
            console.log('生成 .cpt 配置:', config);
            alert('正在生成 .cpt 文件...\\n配置: ' + JSON.stringify(config, null, 2));
        }}
        
        function exportJson() {{
            const data = JSON.stringify(excelData, null, 2);
            const blob = new Blob([data], {{type: 'application/json'}});
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = 'excel_structure.json';
            a.click();
        }}
        
        function applyConfig() {{
            alert('配置已应用');
        }}
        
        // 初始化字段映射
        const fieldMapping = document.getElementById('fieldMapping');
        for (let i = 0; i < 5; i++) {{
            const row = document.createElement('div');
            row.className = 'field-row';
            row.innerHTML = `
                <span>{{get_column_letter(i+1)}}</span>
                <input type="text" placeholder="数据库字段名">
                <select>
                    <option value="string">文本</option>
                    <option value="number">数字</option>
                    <option value="date">日期</option>
                </select>
            `;
            fieldMapping.appendChild(row);
        }}
    </script>
</body>
</html>
''')
        
        return ''.join(html_parts)
    
    def to_summary(self, structure: ExcelStructure) -> Dict[str, Any]:
        """生成结构摘要"""
        return {
            "sheets_count": len(structure.sheets),
            "sheets": [
                {
                    "name": s.name,
                    "rows": s.max_row + 1,
                    "columns": s.max_column + 1,
                    "merged_cells": len(s.merged_ranges),
                    "has_filter": s.filter_range is not None,
                    "styles_count": len(s.styles)
                }
                for s in structure.sheets
            ],
            "defined_names": list(structure.defined_names.keys())
        }


# 测试代码
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <file.xlsx>")
        sys.exit(1)
    
    parser = ExcelParser()
    structure = parser.parse(sys.argv[1])
    
    summary = parser.to_summary(structure)
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    
    # 生成预览 HTML
    html = parser.generate_preview_html(structure)
    output_path = sys.argv[1].replace('.xlsx', '_preview.html')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"\n✅ 预览 HTML 已生成: {output_path}")